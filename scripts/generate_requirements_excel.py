#!/usr/bin/env python3
"""Generate detailed FSM platform requirements Excel workbook."""

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

OUTPUT = "/Users/Vikram/Downloads/knopvvs/docs/FSM_Platform_Requirements.xlsx"

HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
SUBHEADER_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)


def style_header_row(ws, row=1):
    for cell in ws[row]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER


def auto_width(ws, min_width=10, max_width=50):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, min_width), max_width)


def write_sheet(ws, headers, rows, freeze=True):
    ws.append(headers)
    style_header_row(ws)
    for row in rows:
        ws.append(row)
    if freeze:
        ws.freeze_panes = "A2"
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = THIN_BORDER
    auto_width(ws)


# fmt: off
REQUIREMENTS = [
    # ID, Module, Title, Description, Type, Priority, Phase, Source, Owner, System Owner, Integration, AI Level, Acceptance Criteria, Dependencies, Open Questions, Notes
    ["REQ-001", "Platform Foundation", "Multi-tenant organization model", "Each customer company (tenant) has isolated data. All core entities scoped by organization_id. Supports future SaaS commercialization.", "Functional", "P0", "Phase 1", "Architecture", "Platform", "Our Platform", "None", "None", "Tenants cannot access each other's data; RLS or equivalent enforced", "REQ-002", "", "Required even for single client MVP"],
    ["REQ-002", "Authentication", "User authentication and sessions", "Secure login for office staff, dispatchers, technicians, subcontractors. Session management with logout and device tracking.", "Functional", "P0", "Phase 1", "Brief", "Platform", "Our Platform", "Microsoft Entra ID (optional SSO)", "None", "Users authenticate and access only permitted modules per role", "REQ-001", "", "Microsoft SSO natural fit with Graph integrations"],
    ["REQ-003", "Authentication", "Role-based access control", "Roles: Owner, Dispatcher, Technician, Accountant, Subcontractor (limited), Read-only. Permissions per module and action.", "Functional", "P0", "Phase 1", "Architecture", "Platform", "Our Platform", "None", "None", "Each role sees only permitted screens and API operations", "REQ-002", "", ""],
    ["REQ-004", "Authentication", "Microsoft SSO / Entra ID", "Allow login with Microsoft 365 credentials (same tenant as Outlook/OneDrive).", "Functional", "P1", "Phase 1", "Meeting", "Jesper Knopp", "Our Platform", "Microsoft Graph", "None", "User can sign in with M365 account used for Outlook", "REQ-002", "", "Reduces separate password management"],
    ["REQ-005", "Audit", "Activity and audit log", "Record who created/changed entities, sync events, AI decisions, and communications. Immutable audit trail.", "Non-Functional", "P0", "Phase 1", "Architecture", "Platform", "Our Platform", "None", "None", "All critical actions logged with actor, timestamp, before/after", "REQ-001", "", ""],

    ["REQ-010", "CRM", "Customer account model", "Model Account (commercial entity), Contact (people), ServiceLocation (addresses). Support multiple sites per account.", "Functional", "P0", "Phase 1", "Architecture", "Platform", "Our Platform", "Ordrestyring Debtors", "None", "Account can have multiple contacts and service locations", "REQ-011", "", "Maps to Ordrestyring debtor"],
    ["REQ-011", "CRM", "Sync customers from Ordrestyring", "Pull debtors from Ordrestyring API. Store mapping our_account_id ↔ ordrestyring_debtor_id. Periodic sync for updates.", "Functional", "P0", "Phase 1", "Meeting", "Jesper Knopp", "Ordrestyring", "Ordrestyring REST/GraphQL", "None", "Debtors from Ordrestyring appear in CRM within sync interval; mapping stored", "REQ-012", "Confirm API credentials available", ""],
    ["REQ-012", "CRM", "Create new customer in Ordrestyring", "When AI intake identifies new customer and dispatcher approves, create debtor in Ordrestyring and local CRM.", "Functional", "P0", "Phase 2", "Meeting", "Jesper Knopp", "Ordrestyring", "Ordrestyring REST", "Medium", "Approved new customer creates debtor in OS without duplicate on retry", "REQ-011", "Auto-create vs manual approval for new debtors?", ""],
    ["REQ-013", "CRM", "Customer communication preferences", "Store preferred channel (SMS/email/call), language, on-my-way preference (call vs SMS).", "Functional", "P1", "Phase 2", "Client Doc", "Jesper Knopp", "Our Platform", "Twilio, Email", "None", "Preferences used by automated customer messages", "REQ-010", "", ""],
    ["REQ-014", "CRM", "Customer timeline view", "Single view per customer: all jobs, messages, calls, emails, invoices status.", "Functional", "P1", "Phase 3", "Architecture", "Platform", "Our Platform", "Ordrestyring", "None", "Timeline shows linked cases, comms, and sync status", "REQ-010", "", "Salesforce-style UX win"],

    ["REQ-020", "Incoming Orders", "Centralize inbound channels", "Collect inquiries and orders from email, SMS, and requisitions into one central inbox.", "Functional", "P0", "Phase 2", "Client Doc", "Jesper Knopp", "Our Platform", "Outlook, Twilio", "Low", "Email and SMS appear in unified inbox within 2 min of receipt", "REQ-021", "", ""],
    ["REQ-021", "Incoming Orders", "Email integration via Outlook", "Connect shared/service mailbox via Microsoft Graph. Webhook or polling for new messages. Store raw payload immutably.", "Functional", "P0", "Phase 2", "Brief + Meeting", "Jesper Knopp", "Our Platform", "Microsoft Graph", "None", "New emails in connected mailbox create InboundMessage records", "REQ-002", "Shared mailbox vs per-user?", ""],
    ["REQ-022", "Incoming Orders", "SMS integration", "Inbound/outbound SMS via Twilio. Thread linked to customer and work order.", "Functional", "P0", "Phase 2", "Brief", "Jesper Knopp", "Our Platform", "Twilio", "Low", "Inbound SMS creates InboundMessage; outbound logged on case", "REQ-020", "", "Regulatory compliance per country"],
    ["REQ-023", "Incoming Orders", "AI triage of inbound messages", "Classify: service order, quote request, complaint, follow-up, spam, supplier, internal.", "Functional", "P0", "Phase 2", "Architecture", "Platform", "Our Platform", "OpenAI", "High", "Messages routed to correct queue with category and urgency", "REQ-024", "", ""],
    ["REQ-024", "Incoming Orders", "AI structured data extraction", "Extract customer info, address, task type, contact details, problem description from email body and attachments.", "Functional", "P0", "Phase 2", "Client Doc", "Jesper Knopp", "Our Platform", "OpenAI", "High", "Structured JSON output with per-field confidence scores", "REQ-025", "", ""],
    ["REQ-025", "Incoming Orders", "AI PDF and attachment processing", "OCR and layout analysis on PDF attachments. Extract relevant job data.", "Functional", "P0", "Phase 2", "Brief", "Jesper Knopp", "Our Platform", "OpenAI Vision", "High", "PDF attachments processed; extracted fields appear in draft", "REQ-024", "", ""],
    ["REQ-026", "Incoming Orders", "Entity resolution against CRM", "Match extracted data to existing Ordrestyring debtor / local account by name, phone, email, address.", "Functional", "P0", "Phase 2", "Architecture", "Platform", "Our Platform", "Ordrestyring", "High", "Match suggestions shown with confidence; user can confirm or create new", "REQ-011", "", ""],
    ["REQ-027", "Incoming Orders", "Auto-populate draft order (~80%)", "Generate draft work order with majority of required fields filled from available information.", "Functional", "P0", "Phase 2", "Client Doc", "Jesper Knopp", "Our Platform", "OpenAI", "High", "≥80% of defined required fields populated without manual entry (KPI TBD)", "REQ-028", "Define which fields count toward 80%", "Measurable KPI needed"],
    ["REQ-028", "Incoming Orders", "Dispatcher review queue", "Office reviews AI draft before creating official case. Edit fields, approve, reject, or request more info.", "Functional", "P0", "Phase 2", "Brief", "Jesper Knopp", "Our Platform", "None", "Low", "No Ordrestyring case created without explicit approval (configurable auto-approve rules later)", "REQ-029", "", "Human-in-the-loop by default"],
    ["REQ-029", "Incoming Orders", "Duplicate detection", "Detect duplicate inbound messages and prevent duplicate case creation for same request.", "Functional", "P1", "Phase 2", "Architecture", "Platform", "Our Platform", "OpenAI", "Medium", "System warns when similar open case exists for same customer/address", "REQ-028", "", ""],
    ["REQ-030", "Incoming Orders", "Provisional order before requisition", "Support job received before formal requisition arrives days later. Hold provisional state; attach requisition when received.", "Functional", "P1", "Phase 2", "Client Doc", "Jesper Knopp", "Our Platform", "Ordrestyring", "Low", "Provisional case can be created and scheduled; requisition enriches without duplicate", "REQ-028", "Client process rules for billing without requisition?", "Open question from client doc"],

    ["REQ-040", "Service Orders", "Work order entity", "Local work order mirrors Ordrestyring case with enriched metadata, AI provenance, and sync state.", "Functional", "P0", "Phase 1", "Brief + Meeting", "Jesper Knopp", "Our Platform", "Ordrestyring", "None", "Work order stores mapping to OS case ID and sync status", "REQ-041", "", ""],
    ["REQ-041", "Service Orders", "Create case in Ordrestyring", "On approval, POST case to Ordrestyring via API. Idempotent — no duplicates on retry.", "Functional", "P0", "Phase 1", "Meeting", "Jesper Knopp", "Ordrestyring", "Ordrestyring REST/GraphQL", "None", "Approved draft creates exactly one OS case; mapping stored", "REQ-040", "Confirm API access", ""],
    ["REQ-042", "Service Orders", "Bidirectional case sync", "Pull case updates from Ordrestyring (status, dates, manual changes). Push our enriched fields where API allows.", "Functional", "P0", "Phase 1", "Meeting", "Jesper Knopp", "Ordrestyring", "Ordrestyring GraphQL", "None", "Changes in OS reflected in our UI within sync interval; conflict rules documented", "REQ-041", "", ""],
    ["REQ-043", "Service Orders", "Practical access information on order", "Store key box location, door/gate/alarm codes, property access, entrances, parking, on-site contacts, safety notes.", "Functional", "P0", "Phase 1", "Client Doc", "Jesper Knopp", "Our Platform", "Ordrestyring", "None", "All access fields visible on order and mobile before tech arrives", "REQ-040", "", "Appears 3x in client doc — single module"],
    ["REQ-044", "Service Orders", "Work order status state machine", "States: draft, pending_review, scheduled, dispatched, in_progress, on_hold, waiting_for_customer, completed, invoiced, closed, cancelled.", "Functional", "P0", "Phase 1", "Architecture", "Jesper Knopp", "Our Platform", "Ordrestyring", "None", "Every transition logged with actor and timestamp", "REQ-042", "", ""],
    ["REQ-045", "Service Orders", "Waiting for customer status", "When customer cancels and will reschedule, case placed on hold with clear status text.", "Functional", "P1", "Phase 2", "Client Doc", "Jesper Knopp", "Our Platform", "Ordrestyring", "None", "Status visible on order, calendar, and mobile", "REQ-044", "", ""],
    ["REQ-046", "Service Orders", "Incomplete job handling", "If job not completed, original visit record preserved. New planning task for remainder — do not move original.", "Functional", "P0", "Phase 2", "Client Doc", "Jesper Knopp", "Our Platform", "Outlook, Ordrestyring", "None", "New Outlook task created for remainder; original appointment unchanged", "REQ-060", "", ""],

    ["REQ-050", "Quotations", "Quotation intake pipeline", "Inbound quote inquiries (email/SMS) registered as quotation tasks with images, address, contact, description, drawings.", "Functional", "P1", "Phase 4", "Client Doc", "Jesper Knopp", "Our Platform", "OpenAI", "Medium", "Quote inquiries appear in quotation queue separate from service orders", "REQ-020", "", ""],
    ["REQ-051", "Quotations", "Quotation status overview", "Show quotes to prepare, ready for pricing, awaiting customer response, missing information.", "Functional", "P1", "Phase 4", "Client Doc", "Jesper Knopp", "Our Platform", "None", "None", "Dashboard shows quotation pipeline counts and statuses", "REQ-050", "", ""],
    ["REQ-052", "Quotations", "On-site AI voice quoting", "Tech walks site, speaks naturally; AI maps to trade categories (water, drainage, heating, etc.) and generates quote basis.", "Functional", "P2", "Phase 5", "Client Doc", "Jesper Knopp", "Our Platform", "OpenAI", "Very High", "Voice input structured under company quote headings; exportable draft", "REQ-050", "", "Tier 4 — post-MVP"],
    ["REQ-053", "Quotations", "Export quote to Excel/accounting", "Completed quote basis exportable to Excel or accounting software.", "Functional", "P2", "Phase 5", "Client Doc", "Jesper Knopp", "Our Platform", "Ordrestyring", "Medium", "Export produces valid file matching company template", "REQ-052", "Ordrestyring quote API if available", ""],

    ["REQ-060", "Planning & Calendar", "Outlook task on order creation", "When order/case created, auto-create Outlook task in Planning/Unscheduled list via Microsoft Graph. No date/time.", "Functional", "P0", "Phase 2", "Client Doc + Meeting", "Jesper Knopp", "Outlook", "Microsoft Graph", "None", "Task created within 1 min of case approval with customer, address, description, link", "REQ-041", "Confirm Outlook task list names", "Outlook IS the planning UI"],
    ["REQ-061", "Planning & Calendar", "Outlook task content", "Task includes: title, description, customer name/phone, address, job description, notes, attachment links, order number, deep link to our app.", "Functional", "P0", "Phase 2", "Client Doc", "Jesper Knopp", "Outlook", "Microsoft Graph", "None", "All specified fields present in Outlook task body", "REQ-060", "", ""],
    ["REQ-062", "Planning & Calendar", "Drag-to-calendar scheduling", "Dispatcher drags Outlook task to calendar; system detects appointment and updates schedule.", "Functional", "P0", "Phase 2", "Client Doc + Meeting", "Jesper Knopp", "Outlook", "Microsoft Graph", "None", "Calendar event creation updates assignment dates in our system and Ordrestyring case", "REQ-060", "", "Webhook + reconciliation"],
    ["REQ-063", "Planning & Calendar", "Calendar appointment rich info", "Appointments show address, time, customer, contact, key box, codes, parking, access routes, special notes.", "Functional", "P0", "Phase 2", "Client Doc", "Jesper Knopp", "Outlook", "Microsoft Graph", "None", "Practical info visible in Outlook appointment without opening other systems", "REQ-043", "", ""],
    ["REQ-064", "Planning & Calendar", "Status indicators on appointments", "Visual status: scheduled, customer contacted, agreed, ready for execution, needs action.", "Functional", "P1", "Phase 2", "Client Doc", "Jesper Knopp", "Outlook", "Microsoft Graph", "None", "Status visible on calendar appointment (categories/colors)", "REQ-062", "", ""],
    ["REQ-065", "Planning & Calendar", "Google Maps route link", "Clickable address or 'Open Route in Google Maps' button from calendar/task.", "Functional", "P1", "Phase 2", "Client Doc", "Jesper Knopp", "Our Platform", "Google Maps", "None", "Link opens Google Maps with destination address", "REQ-063", "", "Mobile deep link support"],
    ["REQ-066", "Planning & Calendar", "Follow-up planning task for incomplete work", "Create new Outlook planning task when job cannot be completed; preserve original appointment.", "Functional", "P0", "Phase 2", "Client Doc", "Jesper Knopp", "Outlook", "Microsoft Graph", "None", "New unscheduled task created; original event unchanged", "REQ-046", "", ""],

    ["REQ-070", "Notifications", "Appointment confirmation SMS/email", "On schedule, send customer message with day, time window, access needs, whether they must be home.", "Functional", "P1", "Phase 2", "Client Doc", "Jesper Knopp", "Our Platform", "Twilio, Email", "Low", "Message sent with selected channel; logged on case", "REQ-013", "", ""],
    ["REQ-071", "Notifications", "Delay notification one-click", "One-click professional SMS/email when delayed. Pre-filled customer details and template.", "Functional", "P1", "Phase 3", "Client Doc", "Jesper Knopp", "Our Platform", "Twilio, Email", "Low", "Message sent in <3 taps from mobile or calendar", "REQ-013", "", ""],
    ["REQ-072", "Notifications", "On-my-way SMS", "Send 'arriving within 30 minutes' message. Honor customer preference for call vs SMS.", "Functional", "P1", "Phase 3", "Client Doc", "Jesper Knopp", "Our Platform", "Twilio", "Low", "On-my-way message respects communication preference", "REQ-013", "", ""],
    ["REQ-073", "Notifications", "GPS-based delay suggestion", "Optional: suggest delay message based on GPS/drive time vs scheduled time.", "Functional", "P2", "Phase 4", "Client Doc", "Jesper Knopp", "Our Platform", "Google Maps", "Medium", "System alerts when predicted arrival exceeds scheduled window", "REQ-071", "", "Future enhancement"],

    ["REQ-080", "Time Tracking", "GPS-based time suggestion", "Detect arrive/depart at job address via GPS. Generate suggested time entry linked to case.", "Functional", "P0", "Phase 3", "Client Doc", "Jesper Knopp", "Our Platform", "Google Maps", "Medium", "Arrival/departure detected within configurable geofence radius", "REQ-081", "Geofence radius config", ""],
    ["REQ-081", "Time Tracking", "One-click time approval", "Employee approves suggested time with single tap; entry synced to Ordrestyring hours.", "Functional", "P0", "Phase 3", "Client Doc + Meeting", "Jesper Knopp", "Ordrestyring", "Ordrestyring API", "Low", "Approved time appears on OS case as hours entry", "REQ-080", "", ""],
    ["REQ-082", "Time Tracking", "Visit history per case", "Show which cases visited, arrival, departure, duration, previous visits.", "Functional", "P1", "Phase 3", "Client Doc", "Jesper Knopp", "Our Platform", "Ordrestyring", "None", "Visit history visible on case detail and mobile", "REQ-081", "", ""],
    ["REQ-083", "Time Tracking", "Subcontractor time registration", "Subcontractors register hours on assigned tasks; visible to principal.", "Functional", "P1", "Phase 3", "Client Doc", "Jesper Knopp", "Our Platform", "Ordrestyring", "None", "Subcontractor hours appear on case with attribution", "REQ-120", "", ""],

    ["REQ-090", "Photos & Documents", "OneDrive photo storage", "Work photos stored in structured OneDrive folder via Microsoft Graph.", "Functional", "P0", "Phase 3", "Brief + Meeting", "Jesper Knopp", "OneDrive", "Microsoft Graph", "None", "Photos uploaded to correct case folder in OneDrive", "REQ-091", "Shared drive vs per-tech OneDrive?", ""],
    ["REQ-091", "Photos & Documents", "Auto-link photos to case", "Photos automatically linked to correct case/order. Primary: active job on mobile; secondary: GPS heuristic.", "Functional", "P0", "Phase 3", "Client Doc", "Jesper Knopp", "Our Platform", "OneDrive, Ordrestyring", "Medium", "Photo linked to case without manual upload placement in >95% of active-job captures", "REQ-090", "", "Do not rely on EXIF GPS alone"],
    ["REQ-092", "Photos & Documents", "Upload photo to Ordrestyring case", "Async upload of photos/documents to Ordrestyring case via GraphQL uploadCaseDocument.", "Functional", "P0", "Phase 3", "Meeting", "Jesper Knopp", "Ordrestyring", "Ordrestyring GraphQL", "None", "Photo appears as case document in Ordrestyring", "REQ-090", "", ""],
    ["REQ-093", "Photos & Documents", "Photo gallery on order", "Display all linked photos on work order in web and mobile.", "Functional", "P1", "Phase 3", "Client Doc", "Jesper Knopp", "Our Platform", "OneDrive", "None", "All case photos visible with OneDrive links", "REQ-091", "", ""],

    ["REQ-100", "Invoicing", "AI invoice draft generation", "On job completion, AI generates ~90% complete invoice draft from hours, materials, descriptions on case.", "Functional", "P1", "Phase 4", "Client Doc + Meeting", "Jesper Knopp", "Our Platform", "OpenAI, Ordrestyring", "High", "Invoice draft generated with one click; ≥90% line items correct (KPI TBD)", "REQ-101", "Define 90% measurement", ""],
    ["REQ-101", "Invoicing", "Invoice review and edit", "Employee reviews draft, corrects errors, adjusts wording before finalize.", "Functional", "P1", "Phase 4", "Client Doc", "Jesper Knopp", "Our Platform", "None", "Low", "All invoice fields editable before push to Ordrestyring", "REQ-100", "", ""],
    ["REQ-102", "Invoicing", "Push invoice to Ordrestyring", "Submit approved invoice draft to Ordrestyring debtor invoices.", "Functional", "P1", "Phase 4", "Meeting", "Jesper Knopp", "Ordrestyring", "Ordrestyring REST", "None", "Invoice created in OS or exported per API capability", "REQ-101", "Validate invoice API capabilities", ""],

    ["REQ-110", "Forms & QA", "Service report auto-fill", "Service report pre-filled from order: customer, address, contact, order #, date, tech, facility, service agreement.", "Functional", "P1", "Phase 4", "Client Doc", "Jesper Knopp", "Our Platform", "Ordrestyring", "None", "Report fields auto-populated; tech completes checks only", "REQ-111", "", ""],
    ["REQ-111", "Forms & QA", "Service report completion", "Tech records checks, measurements, faults, work performed, parts replaced, remarks, customer signature.", "Functional", "P1", "Phase 4", "Client Doc", "Jesper Knopp", "Our Platform", "OneDrive", "None", "Completed report saved as PDF to case and OneDrive", "REQ-110", "", ""],
    ["REQ-112", "Forms & QA", "Quality assurance report", "QA report from materials list on case. Templates for consistent documentation.", "Functional", "P2", "Phase 4", "Client Doc", "Jesper Knopp", "Our Platform", "Ordrestyring", "None", "QA report generated from case materials with selectable template", "REQ-110", "", "Same forms engine as service report"],

    ["REQ-120", "Subcontractors", "Subcontractor portal access", "Subcontractors view assigned tasks via calendar/app with address, time, case info.", "Functional", "P1", "Phase 3", "Client Doc", "Jesper Knopp", "Our Platform", "None", "None", "Subcontractor sees only assigned tasks", "REQ-003", "", ""],
    ["REQ-121", "Subcontractors", "Subcontractor photos and hours", "Subcontractors upload images and register hours on assigned cases.", "Functional", "P1", "Phase 3", "Client Doc", "Jesper Knopp", "Our Platform", "OneDrive, Ordrestyring", "None", "Uploads linked to correct case; hours visible to principal", "REQ-120", "", ""],
    ["REQ-122", "Subcontractors", "Subcontractor communication", "In-app communication between principal and subcontractor on task.", "Functional", "P2", "Phase 4", "Client Doc", "Jesper Knopp", "Our Platform", "None", "None", "Messages threaded per case with subcontractor", "REQ-120", "", ""],
    ["REQ-123", "Subcontractors", "Materials not entered by subcontractor", "Materials auto-linked to case; subcontractors do not manually register materials.", "Functional", "P1", "Phase 3", "Client Doc", "Jesper Knopp", "Ordrestyring", "Ordrestyring", "None", "Material fields read-only for subcontractor role", "REQ-120", "", ""],

    ["REQ-130", "Follow-up Tasks", "Central follow-up inbox", "Quick capture of small tasks, reminders, follow-ups via type, voice, or photo.", "Functional", "P1", "Phase 4", "Client Doc", "Jesper Knopp", "Our Platform", "OpenAI", "Medium", "User can capture follow-up in <30 sec; appears in central inbox", "REQ-131", "", ""],
    ["REQ-131", "Follow-up Tasks", "Follow-up overview", "Outstanding tasks visible without searching email, SMS, or multiple systems.", "Functional", "P1", "Phase 4", "Client Doc", "Jesper Knopp", "Our Platform", "None", "None", "Inbox shows open follow-ups with source and age", "REQ-130", "", ""],
    ["REQ-132", "Follow-up Tasks", "Miscellaneous uncategorized items", "Catch-all for items with no fixed rule; triage into quote/order/follow-up/spam.", "Functional", "P2", "Phase 4", "Client Doc", "Jesper Knopp", "Our Platform", "OpenAI", "Medium", "Uncategorized items appear in triage queue", "REQ-023", "", "Client explicit catch-all"],

    ["REQ-140", "Voice AI", "Verbal order capture", "Speak order when no paperwork; converted to written order → processing → calendar.", "Functional", "P2", "Phase 5", "Client Doc", "Jesper Knopp", "Our Platform", "OpenAI", "High", "Voice creates draft work order through standard review flow", "REQ-024", "", "Consolidate with voice pipeline"],
    ["REQ-141", "Voice AI", "Professional email from speech", "Speak rough notes; AI produces professional business email. Mobile-first.", "Functional", "P2", "Phase 5", "Client Doc", "Jesper Knopp", "Our Platform", "OpenAI", "High", "Email draft generated with adjustable tone; send via connected mailbox", "REQ-021", "", ""],

    ["REQ-150", "Property Data", "Danish property register lookup", "Enter address; retrieve BBR, OIS, WebLager, utility maps, sewer drawings, heating/water supply.", "Functional", "P2", "Phase 5", "Client Doc", "Jesper Knopp", "Our Platform", "External DK APIs", "Low", "Property data displayed for valid Danish addresses where APIs accessible", "REQ-010", "API access and licensing per register", "Denmark-specific"],
    ["REQ-151", "Field Tools", "Water flow calculator", "Calculate flow, pressure drop, valve setting from pipe dimension, flow, pressure, valve type.", "Functional", "P2", "Phase 5", "Client Doc", "Jesper Knopp", "Our Platform", "None", "None", "Calculation returns expected flow, pressure drop, recommended setting", "REQ-140", "", "Separate product bet"],
    ["REQ-152", "Field Tools", "Smart product assistant", "Photo of product → find manual, translate text, step-by-step guide, YouTube video.", "Functional", "P2", "Phase 5", "Client Doc", "Jesper Knopp", "Our Platform", "OpenAI Vision", "Very High", "Product identified; relevant manual/guide returned", "REQ-153", "", "Duplicate with Translation App"],
    ["REQ-153", "Field Tools", "Translation app", "Photo text translation and product usage explanation.", "Functional", "P2", "Phase 5", "Client Doc", "Jesper Knopp", "Our Platform", "OpenAI Vision", "Very High", "Merged with REQ-152 in implementation", "REQ-152", "", "Consolidate with REQ-152"],
    ["REQ-154", "Field Tools", "Troubleshooting knowledge base", "Internal wiki: faults, symptoms, causes, solutions by category (water, heating, ventilation, drains).", "Functional", "P2", "Phase 5", "Client Doc", "Jesper Knopp", "Our Platform", "OpenAI", "Medium", "Searchable KB with category navigation and contribution workflow", "", "", "Content-heavy"],

    ["REQ-160", "Accounting", "Payment reminder overview", "Show customers with 1st/2nd/3rd reminders, unpaid invoices, contact details, follow-up needed.", "Functional", "P2", "Phase 5", "Client Doc", "Jesper Knopp", "Ordrestyring", "Ordrestyring", "Low", "Reminder status visible without opening accounting system", "REQ-102", "Data source: Ordrestyring vs e-conomic", ""],
    ["REQ-161", "Accounting", "AI polite pre-reminder email", "Generate polite email before formal reminder issued.", "Functional", "P2", "Phase 5", "Client Doc", "Jesper Knopp", "Our Platform", "OpenAI", "Medium", "Draft email generated from invoice context", "REQ-160", "", ""],

    ["REQ-170", "Reporting", "Daily email report", "End-of-day email summarizing where team has been and activity.", "Functional", "P1", "Phase 4", "Client Doc", "Jesper Knopp", "Our Platform", "Email", "Low", "Report delivered daily to configured recipients", "REQ-082", "", ""],
    ["REQ-171", "Reporting", "Operations dashboard", "Dashboard: open drafts, sync status, today's jobs, AI accuracy metrics, quotation pipeline.", "Functional", "P1", "Phase 3", "Brief", "Platform", "Our Platform", "Ordrestyring", "None", "Dashboard loads key metrics without external system navigation", "REQ-040", "", ""],
    ["REQ-172", "Reporting", "AI accuracy metrics", "Track field-level correction rate, time-to-draft-order, invoice draft accuracy over time.", "Functional", "P2", "Phase 4", "Architecture", "Platform", "Our Platform", "None", "Metrics dashboard with 80%/90% KPI tracking", "REQ-027", "", ""],

    ["REQ-180", "Mobile App", "Technician mobile app", "View assigned jobs, practical info, navigate, capture photos, approve time, complete forms.", "Functional", "P0", "Phase 3", "Brief", "Jesper Knopp", "Our Platform", "Ordrestyring, OneDrive", "None", "Tech completes daily workflow from mobile without desktop", "REQ-043", "", "React Native recommended"],
    ["REQ-181", "Mobile App", "Offline mode", "Core job data available offline; sync when connectivity returns.", "Functional", "P1", "Phase 3", "Architecture", "Platform", "Our Platform", "None", "None", "Assigned jobs viewable offline; captures queue for sync", "REQ-180", "", "Basements have no signal"],
    ["REQ-182", "Mobile App", "Offline sync conflict resolution", "Rules for merging offline captures with server state.", "Functional", "P1", "Phase 3", "Architecture", "Platform", "Our Platform", "None", "None", "Conflicts surfaced to user; no silent data loss", "REQ-181", "", ""],

    ["REQ-190", "Integration Platform", "Ordrestyring sync engine", "Queue-based sync with retry, backoff, idempotency, dead letter queue, conflict resolution.", "Functional", "P0", "Phase 1", "Meeting", "Platform", "Our Platform", "Ordrestyring", "None", "Failed syncs retry and surface in admin UI; no duplicate entities", "REQ-041", "", "Core infrastructure"],
    ["REQ-191", "Integration Platform", "Microsoft Graph integration", "Unified adapter for Outlook tasks, calendar events, OneDrive, mailbox.", "Functional", "P0", "Phase 2", "Meeting", "Platform", "Our Platform", "Microsoft Graph", "None", "Graph operations authenticated per tenant with token refresh", "REQ-060", "", ""],
    ["REQ-192", "Integration Platform", "Sync mapping store", "Persist our_id ↔ external_id for Ordrestyring, Outlook, OneDrive per entity type.", "Functional", "P0", "Phase 1", "Architecture", "Platform", "Our Platform", "All", "None", "Every synced entity has retrievable external mapping", "REQ-190", "", ""],
    ["REQ-193", "Integration Platform", "Graceful degradation", "Platform functional when Ordrestyring or Graph temporarily unavailable; sync catches up.", "Non-Functional", "P0", "Phase 1", "Architecture", "Platform", "Our Platform", "All", "None", "Intake and mobile work during outage; sync backlog clears on recovery", "REQ-190", "", ""],

    ["REQ-200", "Workflow Engine", "Event-driven workflow orchestration", "Domain events trigger side effects: notifications, sync, Outlook task creation, AI pipelines.", "Functional", "P0", "Phase 1", "Brief", "Platform", "Our Platform", "All", "None", "Approved order triggers defined chain within 60 seconds", "REQ-190", "BullMQ vs Temporal", ""],
    ["REQ-201", "Workflow Engine", "Configurable business rules", "Per-tenant rules: auto-approve thresholds, notification templates, sync field mapping.", "Functional", "P1", "Phase 3", "Architecture", "Platform", "Our Platform", "None", "None", "Admin can configure rules without code deployment", "REQ-200", "", ""],

    ["REQ-210", "Security", "Tenant data isolation", "Strict isolation between organizations at DB and API layers.", "Non-Functional", "P0", "Phase 1", "Architecture", "Platform", "Our Platform", "None", "None", "Penetration test confirms no cross-tenant data access", "REQ-001", "", ""],
    ["REQ-211", "Security", "Encryption at rest and in transit", "TLS 1.3 in transit; Azure-managed encryption at rest for DB and blobs.", "Non-Functional", "P0", "Phase 1", "Architecture", "Platform", "Azure", "None", "None", "Security audit confirms encryption standards", "", "", ""],
    ["REQ-212", "Security", "Secrets management", "API keys and tokens in Azure Key Vault; never in source code.", "Non-Functional", "P0", "Phase 1", "Architecture", "Platform", "Azure", "None", "None", "No secrets in repository; rotation supported", "", "", ""],
    ["REQ-213", "Security", "AI output validation", "All AI-extracted data validated before DB write; treat as untrusted input.", "Non-Functional", "P0", "Phase 2", "Architecture", "Platform", "Our Platform", "OpenAI", "None", "Malformed AI output rejected with error logged", "REQ-024", "", ""],
    ["REQ-214", "Security", "GDPR compliance", "Data export, erasure, consent for communications, DPA with subprocessors.", "Non-Functional", "P0", "Phase 1", "Architecture", "Platform", "Our Platform", "None", "None", "GDPR checklist completed before production launch", "", "", "EU/Denmark market"],
    ["REQ-215", "Security", "Webhook signature verification", "Verify signatures on Twilio, Graph, and other inbound webhooks.", "Non-Functional", "P0", "Phase 2", "Architecture", "Platform", "Our Platform", "All", "None", "Unsigned webhooks rejected", "", "", ""],

    ["REQ-220", "API Layer", "REST API for platform", "Versioned REST API for web, mobile, and future integrations.", "Functional", "P1", "Phase 2", "Brief", "Platform", "Our Platform", "None", "None", "OpenAPI spec published; auth required on all endpoints", "REQ-002", "", ""],
    ["REQ-221", "API Layer", "Public webhooks outbound", "Tenants can subscribe to events (case.created, job.completed, etc.).", "Functional", "P2", "Phase 5", "Architecture", "Platform", "Our Platform", "None", "None", "Webhook delivery with retry and signing", "REQ-220", "", "Platform play"],

    ["REQ-230", "SaaS Platform", "Subscription and billing", "Tenant subscription tiers, feature flags, usage metering (AI tokens, SMS).", "Functional", "P2", "Phase 5", "Architecture", "Platform", "Our Platform", "Stripe", "None", "Billing blocks access on expired subscription", "REQ-001", "", "Commercial launch"],
    ["REQ-231", "SaaS Platform", "Tenant onboarding wizard", "Connect Ordrestyring, Microsoft 365, Twilio; import customers; first AI draft.", "Functional", "P2", "Phase 4", "Architecture", "Platform", "Our Platform", "All", "None", "New tenant operational within 1 business day", "REQ-190", "", ""],
]

INTEGRATION_MATRIX = [
    ["Entity", "Our Platform", "Ordrestyring", "Outlook", "OneDrive", "Master System", "Sync Direction", "API", "Notes"],
    ["Customer / Debtor", "Account + Contacts", "Debtor", "—", "—", "Ordrestyring", "Bi-directional", "REST / GraphQL", "Create on approval; pull updates"],
    ["Work Order / Case", "WorkOrder", "Case", "—", "—", "Ordrestyring", "Bi-directional", "REST / GraphQL", "Our system enriches; OS is legal record"],
    ["Schedule / Assignment", "Assignment", "Case dates", "Calendar Event", "—", "Outlook (planning UX)", "Outlook → Us → OS", "Microsoft Graph", "Hub model — no direct OS↔Outlook"],
    ["Planning Task", "PlanningTaskRef", "—", "To-Do Task", "—", "Outlook", "Us → Outlook", "Microsoft Graph", "Unscheduled until dragged to calendar"],
    ["Time Entry", "TimeEntry (suggested)", "Hours", "—", "—", "Ordrestyring", "Us → OS", "REST / GraphQL", "After employee approval"],
    ["Photo / Document", "FileMetadata", "Case Document", "—", "File", "OneDrive (blob) + OS (official)", "Us → Both", "Graph + GraphQL", "OneDrive for browse; OS for case record"],
    ["Invoice", "InvoiceDraft", "Debtor Invoice", "—", "—", "Ordrestyring", "Us → OS", "REST", "AI drafts in our UI; finalize in OS"],
    ["Inbound Email", "InboundMessage", "—", "Mailbox", "—", "Our Platform", "Outlook → Us", "Microsoft Graph", "Immutable raw store"],
    ["SMS", "InboundMessage / OutboundMessage", "—", "—", "—", "Our Platform", "Twilio ↔ Us", "Twilio", "Logged on case"],
    ["Quotation (pre-case)", "QuotationTask", "—", "—", "—", "Our Platform", "—", "—", "Becomes case when approved"],
    ["Follow-up Task", "FollowUpTask", "—", "—", "—", "Our Platform", "—", "—", "May convert to case later"],
]

OPEN_QUESTIONS = [
    ["ID", "Question", "Impact", "Owner", "Status", "Decision", "Date"],
    ["OQ-001", "What happens when job is received before formal requisition arrives days later?", "Order state machine, billing rules", "Jesper Knopp", "Open", "", ""],
    ["OQ-002", "OneDrive: shared business library or per-technician personal drive?", "Graph permissions, folder structure", "Jesper Knopp", "Open", "", ""],
    ["OQ-003", "Outlook: single shared planning mailbox or per-dispatcher?", "Graph auth, task ownership", "Jesper Knopp", "Open", "", ""],
    ["OQ-004", "Does client have Ordrestyring API credentials / partner agreement today?", "Phase 1 blocker", "Jesper Knopp", "Open", "", ""],
    ["OQ-005", "Invoice flow: finalize in our UI then push, or review in our UI and complete in Ordrestyring UI?", "Invoice module design", "Jesper Knopp", "Open", "", ""],
    ["OQ-006", "New customers: auto-create debtor in Ordrestyring or always require manual approval?", "AI intake automation level", "Jesper Knopp", "Open", "", ""],
    ["OQ-007", "Which fields count toward the 80% auto-populate KPI for orders?", "AI acceptance criteria", "Jesper Knopp", "Open", "", ""],
    ["OQ-008", "Which line items count toward the 90% invoice draft KPI?", "AI acceptance criteria", "Jesper Knopp", "Open", "", ""],
    ["OQ-009", "Ordrestyring quote API — create quotes in OS or export to Excel only?", "Quotation module", "Jesper Knopp", "Open", "", ""],
    ["OQ-010", "Payment reminder data source: Ordrestyring only or separate accounting (e-conomic)?", "Accounting module scope", "Jesper Knopp", "Open", "", ""],
    ["OQ-011", "Does Jesper want to eventually replace Ordrestyring or augment indefinitely?", "Long-term product strategy", "Jesper Knopp", "Open", "Augment confirmed for v1", ""],
    ["OQ-012", "Auto-approve high-confidence AI drafts for repeat customers?", "Workflow rules", "Jesper Knopp", "Open", "", ""],
]

PHASE_PLAN = [
    ["Phase", "Name", "Duration (est.)", "Requirements", "Deliverable", "Success Criteria"],
    ["Phase 0", "Foundation", "4-6 weeks", "REQ-001 to REQ-005, REQ-190 to REQ-193, REQ-200, REQ-210 to REQ-214", "Multi-tenant SaaS skeleton, Ordrestyring adapter, sync engine, auth", "Tenant created; OS debtor/case sync works; audit log active"],
    ["Phase 1", "Core Ops", "6-8 weeks", "REQ-010 to REQ-014, REQ-040 to REQ-045, REQ-171", "CRM mirror, work orders, case sync, basic dashboard", "Dispatcher views OS cases in our UI with enriched access info"],
    ["Phase 2", "AI Intake + Outlook", "8-10 weeks", "REQ-020 to REQ-030, REQ-060 to REQ-066, REQ-070, REQ-191, REQ-213, REQ-215, REQ-220", "Email/SMS inbox, AI pipeline, review queue, Outlook planning", "Email → reviewed draft → OS case + Outlook task in <15 min median"],
    ["Phase 3", "Field + Media", "8-10 weeks", "REQ-080 to REQ-083, REQ-090 to REQ-093, REQ-071, REQ-072, REQ-120 to REQ-123, REQ-180 to REQ-182", "Mobile app, GPS time, OneDrive photos, subcontractor portal, customer comms", "Tech completes job loop mobile-only; time in OS; photos in OneDrive + OS"],
    ["Phase 4", "Close-out + Reports", "6-8 weeks", "REQ-050, REQ-051, REQ-100 to REQ-102, REQ-110 to REQ-112, REQ-130 to REQ-132, REQ-170, REQ-172, REQ-231", "Invoice AI, forms, follow-up inbox, quotations intake, daily report", "Invoice draft → OS; service report PDF on case"],
    ["Phase 5", "Differentiators", "TBD", "REQ-052, REQ-053, REQ-140 to REQ-154, REQ-160, REQ-161, REQ-221, REQ-230", "Voice quoting, property data, field tools, KB, SaaS billing", "Prioritized per client commercial roadmap"],
]

MODULES = [
    ["#", "Module", "Description", "Phase", "Priority", "Integration Dependencies", "Requirement IDs"],
    ["1", "Platform Foundation", "Multi-tenancy, auth, RBAC, audit", "Phase 0-1", "P0", "Azure, Entra ID", "REQ-001 to REQ-005"],
    ["2", "CRM", "Accounts, contacts, locations, OS debtor sync", "Phase 1", "P0", "Ordrestyring", "REQ-010 to REQ-014"],
    ["3", "Incoming Orders & AI", "Inbox, triage, extraction, review queue", "Phase 2", "P0", "OpenAI, Graph, Twilio", "REQ-020 to REQ-030"],
    ["4", "Service Orders", "Work orders, access info, status, OS case sync", "Phase 1-2", "P0", "Ordrestyring", "REQ-040 to REQ-046"],
    ["5", "Planning & Calendar", "Outlook tasks, scheduling, status, maps", "Phase 2", "P0", "Microsoft Graph, Google Maps", "REQ-060 to REQ-066"],
    ["6", "Notifications", "Customer SMS/email, delay, on-my-way", "Phase 2-3", "P1", "Twilio, Email", "REQ-070 to REQ-073"],
    ["7", "Time Tracking", "GPS suggest, approve, sync hours to OS", "Phase 3", "P0", "Ordrestyring, Google Maps", "REQ-080 to REQ-083"],
    ["8", "Photos & Documents", "OneDrive storage, case linking, OS upload", "Phase 3", "P0", "OneDrive, Ordrestyring", "REQ-090 to REQ-093"],
    ["9", "Invoicing", "AI draft, review, push to Ordrestyring", "Phase 4", "P1", "Ordrestyring, OpenAI", "REQ-100 to REQ-102"],
    ["10", "Forms & QA", "Service reports, QA templates, signatures", "Phase 4", "P1", "OneDrive", "REQ-110 to REQ-112"],
    ["11", "Quotations", "Quote pipeline, on-site voice quoting", "Phase 4-5", "P1-P2", "OpenAI, Ordrestyring", "REQ-050 to REQ-053"],
    ["12", "Follow-up Tasks", "Quick capture inbox, misc triage", "Phase 4", "P1", "OpenAI", "REQ-130 to REQ-132"],
    ["13", "Subcontractors", "Portal, hours, photos, comms", "Phase 3", "P1", "OneDrive, Ordrestyring", "REQ-120 to REQ-123"],
    ["14", "Voice AI", "Verbal orders, professional emails", "Phase 5", "P2", "OpenAI", "REQ-140 to REQ-141"],
    ["15", "Field Tools", "Property data, water calc, product assistant, KB", "Phase 5", "P2", "DK APIs, OpenAI", "REQ-150 to REQ-154"],
    ["16", "Accounting", "Payment reminders, pre-reminder emails", "Phase 5", "P2", "Ordrestyring", "REQ-160 to REQ-161"],
    ["17", "Reporting", "Dashboard, daily report, AI metrics", "Phase 3-4", "P1", "Ordrestyring", "REQ-170 to REQ-172"],
    ["18", "Mobile App", "Tech + subcontractor field app, offline", "Phase 3", "P0", "All field integrations", "REQ-180 to REQ-182"],
    ["19", "Integration Platform", "Sync engine, Graph adapter, mapping, degradation", "Phase 0-2", "P0", "All", "REQ-190 to REQ-193"],
    ["20", "Workflow Engine", "Event orchestration, business rules", "Phase 0-3", "P0", "All", "REQ-200 to REQ-201"],
    ["21", "Security", "Isolation, encryption, GDPR, AI validation", "Phase 0-2", "P0", "Azure", "REQ-210 to REQ-215"],
    ["22", "API Layer", "REST API, webhooks", "Phase 2-5", "P1-P2", "—", "REQ-220 to REQ-221"],
    ["23", "SaaS Platform", "Billing, onboarding, feature flags", "Phase 4-5", "P2", "Stripe", "REQ-230 to REQ-231"],
]

GLOSSARY = [
    ["Term", "Definition"],
    ["FSM", "Field Service Management — software for scheduling and managing field workers"],
    ["Ordrestyring (OS)", "Danish order/work management platform — client's system of record for cases, hours, invoices"],
    ["Case", "Ordrestyring term for a work order / job"],
    ["Debtor", "Ordrestyring term for a customer / billing entity"],
    ["Our Platform", "The AI-powered layer we are building — intake, orchestration, mobile, sync hub"],
    ["InboundMessage", "Immutable record of raw email, SMS, or other inbound communication"],
    ["Draft Work Order", "AI-generated order pending dispatcher review — not yet in Ordrestyring"],
    ["Hub Model", "Our platform sits between inbox, Outlook, OneDrive, and Ordrestyring — no direct OS↔Outlook sync"],
    ["Practical Access Info", "Key box, codes, parking, entrance instructions — field-critical data"],
    ["Human-in-the-loop", "AI proposes; human approves before system-of-record writes"],
    ["GraphQL (OS)", "Ordrestyring GraphQL API at graphql.ordrestyring.dk — cases, uploads, cursor pagination"],
    ["REST v2 (OS)", "Ordrestyring REST API at v2.api.ordrestyring.dk — debtors, cases, hours"],
]
# fmt: on


def main():
    wb = Workbook()

    # Sheet 1: Requirements Master
    ws = wb.active
    ws.title = "Requirements"
    req_headers = [
        "ID", "Module", "Title", "Description", "Type", "Priority", "Phase",
        "Source", "Owner", "System Owner", "Integration", "AI Level",
        "Acceptance Criteria", "Dependencies", "Open Questions", "Notes", "Status",
    ]
    req_rows = [r + ["Draft"] for r in REQUIREMENTS]
    write_sheet(ws, req_headers, req_rows)

    # Sheet 2: Modules
    ws2 = wb.create_sheet("Modules")
    write_sheet(ws2, MODULES[0], MODULES[1:])

    # Sheet 3: Integration Matrix
    ws3 = wb.create_sheet("Integration Matrix")
    write_sheet(ws3, INTEGRATION_MATRIX[0], INTEGRATION_MATRIX[1:])

    # Sheet 4: Phase Plan
    ws4 = wb.create_sheet("Phase Plan")
    write_sheet(ws4, PHASE_PLAN[0], PHASE_PLAN[1:])

    # Sheet 5: Open Questions
    ws5 = wb.create_sheet("Open Questions")
    write_sheet(ws5, OPEN_QUESTIONS[0], OPEN_QUESTIONS[1:])

    # Sheet 6: Glossary
    ws6 = wb.create_sheet("Glossary")
    write_sheet(ws6, GLOSSARY[0], GLOSSARY[1:])

    # Sheet 7: Document Info
    ws7 = wb.create_sheet("Document Info")
    info = [
        ["Field", "Value"],
        ["Document Title", "AI-Powered FSM Platform — Software Requirements"],
        ["Version", "1.0"],
        ["Date", "2026-03-10"],
        ["Client", "Jesper Knopp / Knop PVS (Plumbing)"],
        ["Prepared By", "Technical Architecture Team"],
        ["Status", "Draft — Pending Client Review"],
        ["Sources", "CURSOR_PROJECT_BRIEF.md, Plumbing translation.md, Client Meeting (Ordrestyring + Outlook + OneDrive)"],
        ["Total Requirements", str(len(REQUIREMENTS))],
        ["Integration Model", "Hybrid Augmentation — Ordrestyring (system of record), Outlook (planning UI), OneDrive (files), Our Platform (AI + orchestration hub)"],
        ["Priority Legend", "P0 = Must have for MVP | P1 = High value | P2 = Future / differentiator"],
        ["Phase Legend", "Phase 0-4 = Core product | Phase 5 = Advanced features"],
        ["", ""],
        ["Change Log", ""],
        ["Version", "Date", "Author", "Changes"],
        ["1.0", "2026-03-10", "Architecture", "Initial consolidated requirements from brief, client doc, and meeting"],
    ]
    for row in info:
        ws7.append(row)
    style_header_row(ws7, 1)
    ws7.freeze_panes = "A2"
    auto_width(ws7)

    import os
    os.makedirs(os.path.dirname(OUTPUT), exist_ok=True)
    wb.save(OUTPUT)
    print(f"Created: {OUTPUT}")
    print(f"Requirements: {len(REQUIREMENTS)}")
    print(f"Sheets: {', '.join(wb.sheetnames)}")


if __name__ == "__main__":
    main()
